"""Gera o relatorio Excel (.xlsx) com cadastro (usuarios), Fazendas
(produtos/plantio/aplicacoes) e Manejo das 3 safras (cultura, estoque
rapido, anotacoes) -- exportacao restrita a `ALAN_MAURO_USERNAME`, ver
`admin_exportar` em `app.py`."""
import io

from openpyxl import Workbook
from openpyxl.styles import Font

import fungicida_data
import models

MOMENTO_LABELS = {"ts": "TS", "sulco": "Sulco", "folha": "Folha"}
TIPO_LABELS = {"quimico": "Quimico", "biologico": "Biologico"}
SAFRA_LABELS = dict(models.SAFRAS)


def _write_sheet(wb, title, headers, rows):
    ws = wb.create_sheet(title=title)
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
    for row in rows:
        ws.append(row)
    for col in ws.columns:
        length = max((len(str(c.value)) for c in col if c.value is not None), default=8)
        ws.column_dimensions[col[0].column_letter].width = min(max(length + 2, 10), 50)
    return ws


def _usuarios_rows():
    site_names_by_id = {s["id"]: s["site_name"] for s in models.get_all_sites()}
    rows = []
    for u in models.get_all_users():
        if u["is_admin"]:
            fazendas = "Todas (admin)"
        else:
            fazendas = ", ".join(models.get_user_permitted_site_names(u["id"])) or "-"
        report_ids = models.get_user_report_site_ids(u["id"])
        relatorios = ", ".join(sorted(site_names_by_id[sid] for sid in report_ids if sid in site_names_by_id)) or "-"
        rows.append([
            u["username"], u["email"] or "", models.fmt_telefone_br(u["telefone"]) or "",
            "Sim" if u["is_admin"] else "Nao", fazendas, relatorios,
        ])
    rows.sort(key=lambda r: r[0].lower())
    return rows


def _fazendas_produtos_rows():
    rows = []
    for site_name, buckets in models.get_all_farm_produtos().items():
        for (safra, momento, tipo), linhas in buckets.items():
            if momento not in models.MOMENTOS:
                continue  # momento "geral" (estoque rapido) vai na aba de Manejo
            for linha in linhas:
                rows.append([
                    site_name, SAFRA_LABELS.get(safra, safra), MOMENTO_LABELS.get(momento, momento),
                    TIPO_LABELS.get(tipo, tipo), linha["data_anotacao"], linha["nome"], linha["ingrediente_ativo"],
                ])
    rows.sort(key=lambda r: (r[0].lower(), r[1], r[2], r[3]))
    return rows


def _fazendas_plantio_rows():
    rows = []
    for site_name, por_safra in models.get_all_farm_plantio().items():
        for safra, linhas in por_safra.items():
            for linha in linhas:
                rows.append([
                    site_name, SAFRA_LABELS.get(safra, safra),
                    linha["data_plantio"], linha["talhao"], linha["variedade"], linha["ciclo_dias"],
                ])
    rows.sort(key=lambda r: (r[0].lower(), r[1]))
    return rows


def _fazendas_aplicacoes_rows():
    rows = []
    for site_name, por_safra in models.get_all_farm_aplicacoes().items():
        for safra, linhas in por_safra.items():
            for linha in linhas:
                rows.append([
                    site_name, SAFRA_LABELS.get(safra, safra), linha["data_aplicacao"], linha["talhao"],
                    linha["fungicidas_quimicos"], linha["fungicidas_biologicos"],
                ])
    rows.sort(key=lambda r: (r[0].lower(), r[1]))
    return rows


def _manejo_cultura_rows():
    rows = []
    for (site_name, safra), info in models.get_all_farm_culturas().items():
        rows.append([site_name, SAFRA_LABELS.get(safra, safra), info["cultura"] or "", models.fmt_data_br(info["updated_at"]) or ""])
    rows.sort(key=lambda r: (r[0].lower(), r[1]))
    return rows


def _manejo_estoque_rows():
    rows = []
    for site_name, buckets in models.get_all_farm_produtos().items():
        for (safra, momento, tipo), linhas in buckets.items():
            if momento != models.MOMENTO_ESTOQUE_RAPIDO:
                continue
            for linha in linhas:
                rows.append([
                    site_name, SAFRA_LABELS.get(safra, safra), TIPO_LABELS.get(tipo, tipo),
                    linha["data_anotacao"], linha["nome"],
                ])
    rows.sort(key=lambda r: (r[0].lower(), r[1], r[2]))
    return rows


def _manejo_anotacoes_rows():
    rows = [[site_name, doenca, nota or ""] for (site_name, doenca), nota in models.get_all_recommendation_notes().items() if nota]
    rows.sort(key=lambda r: (r[0].lower(), r[1]))
    return rows


def _add_whatsapp_sheet(wb):
    """Uma aba so' com os dois blocos da tela "Relatorio WhatsApp":
    historico de envios (data/hora, fazenda, destinatario, status) e,
    logo abaixo, quem esta cadastrado pra receber cada fazenda hoje."""
    ws = wb.create_sheet(title="WhatsApp")

    ws.append(["Historico de Envios"])
    ws.cell(row=ws.max_row, column=1).font = Font(bold=True, size=12)
    ws.append(["Data/Hora", "Fazenda", "Destinatario", "Telefone", "Status", "Detalhe"])
    for cell in ws[ws.max_row]:
        cell.font = Font(bold=True)
    for log in models.get_whatsapp_envio_log(limit=1_000_000):
        ws.append([
            models.fmt_data_br(log["criado_em"]) or "",
            log["site_name"], log["destinatario"] or "", models.fmt_telefone_br(log["telefone"]) or "",
            "Enviado" if log["ok"] else "Falha", log["mensagem"] or "",
        ])

    ws.append([])
    ws.append(["Cadastrados para Receber"])
    ws.cell(row=ws.max_row, column=1).font = Font(bold=True, size=12)
    ws.append(["Fazenda", "Destinatario", "Telefone"])
    for cell in ws[ws.max_row]:
        cell.font = Font(bold=True)
    site_names = sorted(s["site_name"] for s in models.get_all_sites())
    for site_name in site_names:
        for r in models.get_site_whatsapp_recipients(site_name):
            ws.append([site_name, r["username"], models.fmt_telefone_br(r["telefone"])])

    for col in ws.columns:
        length = max((len(str(c.value)) for c in col if c.value is not None), default=8)
        ws.column_dimensions[col[0].column_letter].width = min(max(length + 2, 10), 50)


def _add_fungicidas_sheet(wb, culturas_ativas):
    """Uma aba com o mesmo dado da tela "Relatorio Fungicidas": um quimico
    por linha (doenca + ingrediente) e uma coluna por cultura ativa com
    "Sim"/"Nao" conforme o checkbox "Registrado para"."""
    overrides = models.get_all_fungicida_overrides()
    registro_bloqueado = models.get_all_fungicida_registro_bloqueado()
    translations = models.get_all_disease_translations()

    rows = []
    for doenca_en, info in sorted(translations.items(), key=lambda kv: kv[1]["nome_pt"]):
        rec = fungicida_data.get_recomendacao(doenca_en)
        if not rec:
            continue
        grupo = rec["quimicos"]
        n = len(grupo["itens"])
        ordem = models.get_fungicida_ordem(doenca_en, "quimico", n)
        for idx in ordem:
            item = grupo["itens"][idx]
            override = overrides.get((doenca_en, "quimico", idx))
            ingrediente = override["ingrediente"] if override else item["ingrediente"]
            culturas_bloqueadas = registro_bloqueado.get((doenca_en, "quimico", idx), set())
            rows.append([info["nome_pt"], ingrediente] + [
                "Nao" if c in culturas_bloqueadas else "Sim" for c in culturas_ativas
            ])

    _write_sheet(wb, "Fungicidas", ["Doenca", "Quimico"] + culturas_ativas, rows)


def build_workbook():
    wb = Workbook()
    wb.remove(wb.active)

    _write_sheet(wb, "Usuarios", ["Usuario", "Email", "Telefone", "Admin", "Fazendas liberadas", "Recebe relatorio"], _usuarios_rows())
    _write_sheet(wb, "Fazendas - Produtos", ["Fazenda", "Safra", "Momento", "Tipo", "Data/Anotacao", "Nome do produto", "Ingrediente ativo"], _fazendas_produtos_rows())
    _write_sheet(wb, "Fazendas - Plantio", ["Fazenda", "Safra", "Data plantio", "Talhao", "Variedade", "Ciclo (dias)"], _fazendas_plantio_rows())
    _write_sheet(wb, "Fazendas - Aplicacoes", ["Fazenda", "Safra", "Data aplicacao", "Talhao", "Fungicidas quimicos", "Fungicidas biologicos"], _fazendas_aplicacoes_rows())
    _write_sheet(wb, "Manejo - Cultura", ["Fazenda", "Safra", "Cultura", "Atualizado em"], _manejo_cultura_rows())
    _write_sheet(wb, "Manejo - Estoque rapido", ["Fazenda", "Safra", "Tipo", "Data/Anotacao", "Nome do produto"], _manejo_estoque_rows())
    _write_sheet(wb, "Manejo - Anotacoes", ["Fazenda", "Doenca", "Nota"], _manejo_anotacoes_rows())
    _add_whatsapp_sheet(wb)
    _add_fungicidas_sheet(wb, models.get_culturas_ativas())

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer
